#!/usr/bin/env python3
"""
find_leads: build lead lists of local businesses ANYWHERE IN THE WORLD using the
official Google Places API (New). For every run it produces TWO sheets:

    * businesses with NO website   (your prospects for selling a website)
    * businesses WITH a website    (for comparison / other offers)

WHY THIS APPROACH
-----------------
Scraping Google Maps directly violates Google's Terms of Service and breaks
constantly. The Places API (New) legally returns each business's website field
(`websiteUri`), so we can split businesses into "has a site" vs "no site".

SETUP
-----
1. Create a Google Cloud project: https://console.cloud.google.com/
2. Enable **Places API (New)** and enable **billing** (Google gives a recurring
   free credit; check current pricing before large runs).
3. Create an API key (restrict it to the Places API).
4. Export it:   export GOOGLE_MAPS_API_KEY="your_key"

USAGE
-----
    pip install -r requirements.txt

    # WORLDWIDE - type any places you want (multiple, no limit)
    python find_leads.py \
        --locations "Paris, France" "New York, USA" "Dubai, UAE" \
        --categories "beauty salon" "dental clinic" "real estate agency"

    # bias results toward a country (ISO 3166-1 code), optional
    python find_leads.py --locations "London" --region GB --categories "cafe"

    # India presets still work (cities / states / everything)
    python find_leads.py --states Assam
    python find_leads.py --cities Guwahati Shillong --categories "gym"
    python find_leads.py --all-india

    # verify the pipeline with no API key (uses a built-in mock response)
    python find_leads.py --self-test

OUTPUT
------
    gmb-leads/output/<prefix>_no_website.csv   (+ .xlsx if openpyxl installed)
    gmb-leads/output/<prefix>_with_website.csv (+ .xlsx)
Columns: business_name, category, location, city, state, phone, address,
         website, maps_url, rating, reviews, business_status, place_id
"""

from __future__ import annotations

import os
import csv
import sys
import json
import time
import argparse
from datetime import datetime

import requests

PLACES_SEARCH_URL = "https://places.googleapis.com/v1/places:searchText"

# Only request the fields we need (keeps cost down and response small).
FIELD_MASK = ",".join([
    "places.id",
    "places.displayName",
    "places.formattedAddress",
    "places.websiteUri",
    "places.nationalPhoneNumber",
    "places.internationalPhoneNumber",
    "places.rating",
    "places.userRatingCount",
    "places.googleMapsUri",
    "places.businessStatus",
    "places.primaryTypeDisplayName",
    "nextPageToken",
])

HERE = os.path.dirname(os.path.abspath(__file__))
PRESETS_PATH = os.path.join(HERE, "presets.json")
OUT_DIR = os.path.join(HERE, "output")

COLUMNS = ["business_name", "category", "location", "city", "state", "phone",
           "address", "website", "maps_url", "rating", "reviews",
           "business_status", "place_id"]


def load_presets() -> dict:
    with open(PRESETS_PATH, encoding="utf-8") as f:
        return json.load(f)


def search_text(api_key: str, query: str, region_code: str | None = None,
                max_pages: int = 3, timeout: int = 30) -> list[dict]:
    """Call Places API (New) Text Search, following pagination up to max_pages.

    region_code is OPTIONAL: pass an ISO 3166-1 code (e.g. "GB", "US", "IN") to
    bias results toward a country, or leave it None for worldwide queries where
    the location text itself disambiguates (e.g. "Paris, France").
    """
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": FIELD_MASK,
    }
    places: list[dict] = []
    page_token = None
    for page in range(max_pages):
        body: dict = {"textQuery": query, "pageSize": 20}
        if region_code:
            body["regionCode"] = region_code
        if page_token:
            body["pageToken"] = page_token
            time.sleep(2)  # token needs a moment to become valid
        for attempt in range(4):
            try:
                resp = requests.post(PLACES_SEARCH_URL, headers=headers,
                                     data=json.dumps(body), timeout=timeout)
            except requests.RequestException as exc:
                print(f"    [network] {exc} (retry {attempt + 1})")
                time.sleep(1 + attempt)
                continue
            if resp.status_code == 429 or resp.status_code >= 500:
                wait = 2 ** attempt
                print(f"    [http {resp.status_code}] backing off {wait}s")
                time.sleep(wait)
                continue
            if resp.status_code != 200:
                print(f"    [http {resp.status_code}] {resp.text[:200]}")
                return places
            data = resp.json()
            places.extend(data.get("places", []))
            page_token = data.get("nextPageToken")
            break
        else:
            break
        if not page_token:
            break
    return places


def extract_rows(places: list[dict], location: str = "", category: str = "",
                 city: str = "", state: str = "") -> list[dict]:
    """Pure function: turn a Places API response list into normalized rows.

    `location` is the free-form place that was searched (e.g. "Paris, France").
    `city`/`state` stay available for the India presets but may be empty for
    worldwide searches.
    """
    rows = []
    for p in places:
        website = p.get("websiteUri", "") or ""
        rows.append({
            "business_name": (p.get("displayName") or {}).get("text", ""),
            "category": category,
            "location": location,
            "city": city,
            "state": state,
            "phone": p.get("nationalPhoneNumber") or p.get("internationalPhoneNumber", ""),
            "address": p.get("formattedAddress", ""),
            "website": website,
            "maps_url": p.get("googleMapsUri", ""),
            "rating": p.get("rating", ""),
            "reviews": p.get("userRatingCount", ""),
            "business_status": p.get("businessStatus", ""),
            "place_id": p.get("id", ""),
            "_website": website,
        })
    return rows


def split_leads(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """Split rows into (no_website, with_website)."""
    no_website = [r for r in rows if not r.get("_website")]
    with_website = [r for r in rows if r.get("_website")]
    return no_website, with_website


def write_csv(path: str, rows: list[dict]) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in COLUMNS})


def write_xlsx(path: str, rows: list[dict], title: str = "Leads",
               header_color: str = "075E54") -> bool:
    """Write a real .xlsx spreadsheet if openpyxl is available; else skip."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append([c.replace("_", " ").title() for c in COLUMNS])
    header_fill = PatternFill("solid", fgColor=header_color)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for r in rows:
        ws.append([r.get(k, "") for k in COLUMNS])
    widths = {"business_name": 32, "category": 20, "location": 24, "city": 14,
              "state": 18, "phone": 16, "address": 50, "website": 34,
              "maps_url": 38, "rating": 8, "reviews": 9,
              "business_status": 16, "place_id": 30}
    for i, c in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 18)
    ws.freeze_panes = "A2"
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return True


def resolve_targets(presets: dict, args) -> list[tuple[str, str, str]]:
    """Return (location, city, state) triples based on CLI filters.

    - --locations: free-form worldwide places, used verbatim.
    - --cities / --states / --all-india: India presets, expanded into
      "City, State, India" location strings.
    """
    # Worldwide free-form locations take priority.
    if args.locations:
        return [(loc.strip(), "", "") for loc in args.locations if loc.strip()]

    states = presets["states"]
    triples: list[tuple[str, str, str]] = []
    if args.cities:
        want = {c.lower() for c in args.cities}
        for state, cities in states.items():
            for city in cities:
                if city.lower() in want:
                    triples.append((f"{city}, {state}, India", city, state))
    elif args.states:
        want = {s.lower() for s in args.states}
        for state, cities in states.items():
            if state.lower() in want:
                triples.extend((f"{city}, {state}, India", city, state) for city in cities)
    elif args.all_india:
        for state, cities in states.items():
            triples.extend((f"{city}, {state}, India", city, state) for city in cities)
    # else: no selection -> empty; main() warns (prevents accidental huge runs)
    return triples


def run_self_test() -> int:
    """Exercise extract/split/write with a mock response (no API key needed)."""
    mock = {
        "places": [
            {"id": "p1", "displayName": {"text": "Hill View Salon"},
             "formattedAddress": "Rue de Rivoli, Paris", "nationalPhoneNumber": "01 23 45 67 89",
             "rating": 4.3, "userRatingCount": 52, "googleMapsUri": "https://maps.google.com/?cid=1",
             "businessStatus": "OPERATIONAL", "websiteUri": ""},
            {"id": "p2", "displayName": {"text": "Seine Dental Care"},
             "formattedAddress": "Avenue Montaigne, Paris", "nationalPhoneNumber": "01 98 76 54 32",
             "rating": 4.7, "userRatingCount": 120, "googleMapsUri": "https://maps.google.com/?cid=2",
             "businessStatus": "OPERATIONAL", "websiteUri": "https://example-dental.fr"},
        ]
    }
    rows = extract_rows(mock["places"], location="Paris, France", category="beauty salon")
    no_website, with_website = split_leads(rows)
    assert len(rows) == 2, "row count wrong"
    assert len(no_website) == 1 and len(with_website) == 1, "split logic broken"
    assert no_website[0]["business_name"] == "Hill View Salon"
    assert with_website[0]["website"] == "https://example-dental.fr"
    no_csv = os.path.join(OUT_DIR, "self_test_no_website.csv")
    yes_csv = os.path.join(OUT_DIR, "self_test_with_website.csv")
    write_csv(no_csv, no_website)
    write_csv(yes_csv, with_website)
    xlsx_ok = write_xlsx(os.path.join(OUT_DIR, "self_test_no_website.xlsx"),
                         no_website, title="No website", header_color="128C7E")
    print("SELF-TEST PASSED")
    print(f"  parsed {len(rows)} places -> {len(no_website)} no-website / "
          f"{len(with_website)} with-website")
    print(f"  wrote {no_csv}")
    print(f"  wrote {yes_csv}")
    print(f"  xlsx written: {xlsx_ok}")
    return 0


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser(
        description="Find local businesses worldwide (with / without a website) via Places API (New).")
    ap.add_argument("--locations", nargs="*",
                    help='free-form places anywhere in the world, e.g. "Paris, France" "Tokyo, Japan"')
    ap.add_argument("--region", default=None,
                    help="optional ISO 3166-1 country code to bias results (e.g. GB, US, IN)")
    ap.add_argument("--cities", nargs="*", help="India presets: limit to these cities")
    ap.add_argument("--states", nargs="*", help="India presets: limit to these states")
    ap.add_argument("--all-india", action="store_true",
                    help="India presets: scan EVERY city in presets.json (large run)")
    ap.add_argument("--categories", nargs="*", help="search keywords / business types")
    ap.add_argument("--max-pages", type=int, default=3, help="pages per query (20 results each, max 3)")
    ap.add_argument("--out", default=os.path.join(OUT_DIR, "worldwide_leads"),
                    help="output path prefix (no extension)")
    ap.add_argument("--api-key", default=os.environ.get("GOOGLE_MAPS_API_KEY"))
    ap.add_argument("--self-test", action="store_true", help="run offline pipeline test (no key needed)")
    # kept for backward compatibility; both sheets are always written now.
    ap.add_argument("--include-with-website", action="store_true",
                    help="(deprecated) both with/without-website sheets are always written")
    args = ap.parse_args(argv)

    if args.self_test:
        return run_self_test()

    if not args.api_key:
        print("ERROR: no API key. Set GOOGLE_MAPS_API_KEY or pass --api-key.\n"
              "       (Run with --self-test to verify the tool works without a key.)")
        return 2

    presets = load_presets()
    categories = args.categories or presets["categories"]
    targets = resolve_targets(presets, args)

    # region: explicit flag wins; for India presets default to the preset region.
    region = args.region
    if region is None and not args.locations:
        region = presets.get("region_code")

    if not targets:
        print("Nothing to search. Specify one of:\n"
              '  --locations "Paris, France" "New York, USA"   (anywhere in the world)\n'
              "  --cities Guwahati Shillong                      (India preset cities)\n"
              "  --states Assam Kerala                           (India preset states)\n"
              "  --all-india                                     (every India preset city)\n")
        return 2

    print(f"Targets: {len(targets)} location(s) x {len(categories)} categories "
          f"= {len(targets) * len(categories)} queries"
          + (f"  [region bias: {region}]" if region else "  [worldwide]"))
    print("(this can take a while and consumes API quota)\n")

    seen: set[str] = set()
    all_rows: list[dict] = []
    for location, city, state in targets:
        for category in categories:
            query = f"{category} in {location}"
            print(f"  - {query}")
            places = search_text(args.api_key, query, region_code=region,
                                 max_pages=args.max_pages)
            for row in extract_rows(places, location=location, category=category,
                                    city=city, state=state):
                pid = row["place_id"]
                if pid and pid in seen:
                    continue
                if pid:
                    seen.add(pid)
                all_rows.append(row)

    no_website, with_website = split_leads(all_rows)
    print(f"\nCollected {len(all_rows)} unique businesses; "
          f"{len(no_website)} have NO website, {len(with_website)} have a website.")

    no_csv = args.out + "_no_website.csv"
    yes_csv = args.out + "_with_website.csv"
    write_csv(no_csv, no_website)
    write_csv(yes_csv, with_website)
    print(f"Wrote {no_csv}")
    print(f"Wrote {yes_csv}")

    xlsx_no = write_xlsx(args.out + "_no_website.xlsx", no_website,
                         title="No website", header_color="128C7E")
    xlsx_yes = write_xlsx(args.out + "_with_website.xlsx", with_website,
                          title="With website", header_color="075E54")
    if xlsx_no and xlsx_yes:
        print(f"Wrote {args.out}_no_website.xlsx and {args.out}_with_website.xlsx")
    else:
        print("(install openpyxl for .xlsx spreadsheets: pip install openpyxl)")

    print(f"\nDone {datetime.utcnow():%Y-%m-%d %H:%M UTC}. "
          f"{len(no_website)} no-website leads ready for outreach.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
